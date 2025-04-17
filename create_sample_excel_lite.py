import openpyxl
from openpyxl.styles import Font, Alignment

def create_sample_excel():
    # 새 워크북 생성
    wb = openpyxl.Workbook()
    
    # 첫 번째 시트 - 직원 정보
    ws_employee = wb.active
    ws_employee.title = "직원정보"
    
    # 헤더 추가
    headers = ['사원번호', '이름', '부서', '입사일', '급여']
    for col, header in enumerate(headers, 1):
        cell = ws_employee.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터 추가
    employee_data = [
        [1001, '김철수', '인사팀', '2020-01-15', 3500000],
        [1002, '이영희', '개발팀', '2018-03-22', 4200000],
        [1003, '박민수', '마케팅', '2021-05-10', 3800000],
        [1004, '최지은', '영업팀', '2019-11-05', 3900000],
        [1005, '정준호', '개발팀', '2022-02-28', 4000000]
    ]
    
    for row_idx, row_data in enumerate(employee_data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws_employee.cell(row=row_idx, column=col_idx).value = cell_value
    
    # 두 번째 시트 - 판매 데이터
    ws_sales = wb.create_sheet(title="판매데이터")
    
    # 헤더 추가
    headers = ['주문번호', '제품명', '가격', '판매량', '판매일']
    for col, header in enumerate(headers, 1):
        cell = ws_sales.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터 추가
    sales_data = [
        ['S001', '노트북', 1200000, 3, '2023-01-15'],
        ['S002', '모니터', 350000, 5, '2023-01-20'],
        ['S003', '키보드', 55000, 10, '2023-02-05'],
        ['S004', '마우스', 35000, 15, '2023-02-10'],
        ['S005', '스피커', 80000, 8, '2023-03-15'],
        ['S006', '헤드폰', 120000, 12, '2023-03-22'],
        ['S007', '태블릿', 800000, 4, '2023-04-10'],
        ['S008', '프린터', 250000, 2, '2023-04-25']
    ]
    
    for row_idx, row_data in enumerate(sales_data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws_sales.cell(row=row_idx, column=col_idx).value = cell_value
    
    # 세 번째 시트 - 재고 정보
    ws_inventory = wb.create_sheet(title="재고정보")
    
    # 헤더 추가
    headers = ['제품코드', '제품명', '카테고리', '공급업체', '현재고', '재주문수량']
    for col, header in enumerate(headers, 1):
        cell = ws_inventory.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터 추가
    inventory_data = [
        ['P001', '노트북', '컴퓨터', '삼성전자', 15, 5],
        ['P002', '모니터', '주변기기', 'LG전자', 23, 10],
        ['P003', '키보드', '주변기기', '로지텍', 50, 20],
        ['P004', '마우스', '주변기기', '로지텍', 45, 20],
        ['P005', '스피커', '오디오', '보스', 12, 5],
        ['P006', '헤드폰', '오디오', '소니', 18, 8],
        ['P007', '태블릿', '모바일', '애플', 10, 5],
        ['P008', '프린터', '주변기기', 'HP', 5, 3]
    ]
    
    for row_idx, row_data in enumerate(inventory_data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws_inventory.cell(row=row_idx, column=col_idx).value = cell_value
    
    # 네 번째 시트 - 거래처 정보
    ws_supplier = wb.create_sheet(title="거래처정보")
    
    # 헤더 추가
    headers = ['업체번호', '업체명', '연락처', '담당자', '이메일', '주소']
    for col, header in enumerate(headers, 1):
        cell = ws_supplier.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터 추가
    supplier_data = [
        ['V001', '삼성전자', '02-123-4567', '홍길동', 'samsung@example.com', '서울시 강남구'],
        ['V002', 'LG전자', '02-234-5678', '이순신', 'lg@example.com', '서울시 영등포구'],
        ['V003', '로지텍', '031-345-6789', '김유신', 'logitech@example.com', '경기도 성남시'],
        ['V004', '소니', '02-456-7890', '강감찬', 'sony@example.com', '서울시 강서구'],
        ['V005', '애플', '02-567-8901', '세종대왕', 'apple@example.com', '서울시 중구']
    ]
    
    for row_idx, row_data in enumerate(supplier_data, 2):
        for col_idx, cell_value in enumerate(row_data, 1):
            ws_supplier.cell(row=row_idx, column=col_idx).value = cell_value
    
    # 컬럼 너비 조정
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for column_cells in ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            adjusted_width = (length + 2)
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
    
    # 엑셀 파일 저장
    wb.save('sample_data.xlsx')
    print("샘플 엑셀 파일 'sample_data.xlsx'가 생성되었습니다.")

if __name__ == "__main__":
    create_sample_excel() 